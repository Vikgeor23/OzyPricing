"""Tests for XLSX product import (legacy and extended templates)."""

import io
import unittest

from openpyxl import Workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from app.database import Base
from app.models import Product
from app.services.product_import import build_template_workbook_bytes, import_products_from_xlsx


def _workbook_bytes(headers: list[str], rows: list[list]) -> bytes:
    wb = Workbook()
    ws = wb.active
    for col, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=h)
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class ProductImportTests(unittest.TestCase):
    def setUp(self) -> None:
        self.engine = create_engine(
            "sqlite:///:memory:",
            connect_args={"check_same_thread": False},
            poolclass=StaticPool,
        )
        Base.metadata.create_all(self.engine, tables=[Product.__table__])
        self.Session = sessionmaker(bind=self.engine, autocommit=False, autoflush=False)

    def test_legacy_template_imports(self) -> None:
        headers = ["sku", "ean", "brand", "name", "category", "manufacturer_code", "own_price"]
        data = _workbook_bytes(headers, [["LEG-1", "5901234123457", "ACME", "Legacy Product", "Phones", "M1", 99.9]])
        db = self.Session()
        try:
            summary = import_products_from_xlsx(db, data)
            self.assertEqual(summary.imported_rows, 1)
            row = db.query(Product).filter(Product.sku == "LEG-1").one()
            self.assertEqual(row.name, "Legacy Product")
            self.assertEqual(row.ean, "5901234123457")
        finally:
            db.close()

    def test_minimal_template_sku_name_only(self) -> None:
        data = _workbook_bytes(["sku", "name"], [["MIN-1", "Minimal Product"]])
        db = self.Session()
        try:
            summary = import_products_from_xlsx(db, data)
            self.assertEqual(summary.imported_rows, 1)
        finally:
            db.close()

    def test_extended_template_with_model(self) -> None:
        headers = ["sku", "name", "model", "ean", "color", "storage"]
        data = _workbook_bytes(
            headers,
            [["EXT-1", "Phone X", "PX-100", "5901234123457", "Black", "128GB"]],
        )
        db = self.Session()
        try:
            summary = import_products_from_xlsx(db, data)
            self.assertEqual(summary.imported_rows, 1)
            row = db.query(Product).filter(Product.sku == "EXT-1").one()
            self.assertEqual(row.model, "PX-100")
            self.assertEqual(row.color, "Black")
        finally:
            db.close()

    def test_build_template_has_required_columns(self) -> None:
        raw = build_template_workbook_bytes()
        summary = import_products_from_xlsx(self.Session(), raw)
        self.assertEqual(summary.total_rows, 0)


if __name__ == "__main__":
    unittest.main()
