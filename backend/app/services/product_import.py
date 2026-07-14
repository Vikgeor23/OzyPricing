"""XLSX product catalog import (synchronous)."""

from __future__ import annotations

import io
import re
from decimal import Decimal, InvalidOperation
from collections.abc import Callable
from typing import Any

from openpyxl import Workbook, load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models import ImportBatch, Product
from app.schemas.product import ProductCreate, ProductUpdate
from app.schemas.product_import import ProductImportErrorItem, ProductImportSummary
from app.services import product_service

# Mandatory columns: EAN + manufacturer code + product name + own price. A sku
# column is accepted but optional (falls back to supplier_sku, then EAN).
MIN_REQUIRED_HEADERS = frozenset({"ean", "manufacturer_code", "name", "own_price"})

# Legacy template (fully supported)
LEGACY_OPTIONAL_HEADERS = frozenset(
    {
        "ean",
        "brand",
        "category",
        "manufacturer_code",
        "own_price",
    },
)

# Extended optional columns
EXTENDED_OPTIONAL_HEADERS = frozenset(
    {
        "model",
        "product_url",
        "image_url",
        "description",
        "variant",
        "color",
        "size",
        "storage",
        "memory",
        "supplier_sku",
    },
)

# Mandatory first, then informative optional columns.
TEMPLATE_HEADERS = [
    "ean",
    "manufacturer_code",
    "name",
    "own_price",
    "sku",
    "brand",
    "category",
    "model",
    "color",
    "size",
    "storage",
    "supplier_sku",
    "product_url",
]

TEMPLATE_EXAMPLE_ROW = [
    "5702016912920",
    "42130",
    "LEGO Technic 42130 - BMW M 1000 RR",
    "259.99",
    "LT-42130",
    "LEGO",
    "Конструктори",
    "Technic",
    "",
    "",
    "",
    "SUP-000123",
    "https://example.com/product/42130",
]


def _norm_header(cell: str | None) -> str:
    if cell is None:
        return ""
    return re.sub(r"\s+", "_", str(cell).strip().lower())


_PLACEHOLDER_VALUES = frozenset({"-", "--", "—", "–", ".", "n/a", "na", "none", "null", "x"})


def _clean(value: str | None) -> str | None:
    if value is None:
        return None
    return None if value.strip().lower() in _PLACEHOLDER_VALUES else value


def _norm_cell(v: Any) -> str | None:
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        return _clean(s) if s else None
    s = str(v).strip()
    return _clean(s) if s else None


def _ean_to_str(v: Any) -> str | None:
    if v is None or (isinstance(v, str) and not v.strip()):
        return None
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        if isinstance(v, float) and v.is_integer():
            return str(int(v))
        return str(v).strip()
    s = str(v).strip()
    return _clean(s) if s else None


def _code_str(v: Any) -> str | None:
    if v is None or (isinstance(v, str) and not v.strip()):
        return None
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        if isinstance(v, float) and v.is_integer():
            return str(int(v))
        return str(v).strip()
    return _clean(str(v).strip())


def _parse_own_price(v: Any) -> Decimal | None:
    if v is None or (isinstance(v, str) and not v.strip()):
        return None
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return Decimal(str(v))
    s = str(v).strip().replace(",", ".")
    if not s:
        return None
    try:
        return Decimal(s)
    except InvalidOperation:
        return None


def _build_header_map(header_row: list[Any]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for idx, cell in enumerate(header_row):
        key = _norm_header(cell)
        if not key:
            continue
        mapping[key] = idx
    return mapping


def _optional_cell(row: tuple[Any, ...] | None, col_map: dict[str, int], name: str) -> Any:
    if name not in col_map:
        return None
    idx = col_map[name]
    if row is None or idx >= len(row):
        return None
    return row[idx]


def import_products_from_xlsx(
    db: Session,
    file_bytes: bytes,
    *,
    filename: str = "catalog.xlsx",
    progress_callback: Callable[[int, int, str], None] | None = None,
) -> ProductImportSummary:
    """Parse workbook and upsert products by SKU under a new import batch."""

    errors: list[ProductImportErrorItem] = []
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    try:
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            return ProductImportSummary(
                total_rows=0,
                imported_rows=0,
                skipped_rows=0,
                errors=[ProductImportErrorItem(row=1, message="Worksheet is empty")],
            )

        col_map = _build_header_map(list(header_row))
        missing = sorted(MIN_REQUIRED_HEADERS - set(col_map.keys()))
        if missing:
            return ProductImportSummary(
                total_rows=0,
                imported_rows=0,
                skipped_rows=0,
                errors=[
                    ProductImportErrorItem(
                        row=1,
                        message=f"Missing required columns: {', '.join(missing)}",
                    ),
                ],
            )

        batch = ImportBatch(filename=filename or "catalog.xlsx")
        db.add(batch)
        db.flush()

        total_rows = 0
        imported_rows = 0

        # Bulk import: collect valid payloads first, then upsert in chunks.
        # Per-row SELECT + COMMIT makes large files (50k+ rows) take minutes
        # and time out behind the proxy.
        payloads: list[ProductCreate] = []

        for excel_row_idx, row in enumerate(rows_iter, start=2):
            total_rows += 1
            if row is None or all(
                c is None or (isinstance(c, str) and not c.strip()) for c in row
            ):
                errors.append(
                    ProductImportErrorItem(row=excel_row_idx, message="Empty row skipped"),
                )
                continue

            name = _norm_cell(_optional_cell(row, col_map, "name"))
            ean = _ean_to_str(_optional_cell(row, col_map, "ean"))
            manufacturer_code = _code_str(_optional_cell(row, col_map, "manufacturer_code"))
            supplier_sku = _code_str(_optional_cell(row, col_map, "supplier_sku"))
            sku = _norm_cell(_optional_cell(row, col_map, "sku")) or supplier_sku or ean

            missing_parts: list[str] = []
            if not ean:
                missing_parts.append("ean is required")
            if not manufacturer_code:
                missing_parts.append("manufacturer_code is required")
            if not name:
                missing_parts.append("name is required")
            if missing_parts:
                errors.append(
                    ProductImportErrorItem(row=excel_row_idx, message="; ".join(missing_parts)),
                )
                continue

            own_cell = _optional_cell(row, col_map, "own_price")
            own_price = _parse_own_price(own_cell)
            if own_price is None:
                errors.append(
                    ProductImportErrorItem(
                        row=excel_row_idx,
                        message=(
                            "own_price is required"
                            if own_cell is None or (isinstance(own_cell, str) and not own_cell.strip())
                            else f"Invalid own_price: {own_cell!r}"
                        ),
                    ),
                )
                continue

            payloads.append(
                ProductCreate(
                    sku=sku,
                    name=name,
                    ean=ean,
                    brand=_norm_cell(_optional_cell(row, col_map, "brand")),
                    category=_norm_cell(_optional_cell(row, col_map, "category")),
                    manufacturer_code=manufacturer_code,
                    model=_code_str(_optional_cell(row, col_map, "model")),
                    own_price=own_price,
                    product_url=_norm_cell(_optional_cell(row, col_map, "product_url")),
                    image_url=_norm_cell(_optional_cell(row, col_map, "image_url")),
                    description=_norm_cell(_optional_cell(row, col_map, "description")),
                    variant=_norm_cell(_optional_cell(row, col_map, "variant")),
                    color=_norm_cell(_optional_cell(row, col_map, "color")),
                    size=_norm_cell(_optional_cell(row, col_map, "size")),
                    storage=_norm_cell(_optional_cell(row, col_map, "storage")),
                    memory=_norm_cell(_optional_cell(row, col_map, "memory")),
                    supplier_sku=supplier_sku,
                ),
            )

        # Preload existing products for all skus in this file (chunked IN).
        skus = [p.sku for p in payloads]
        existing_by_sku: dict[str, Product] = {}
        for i in range(0, len(skus), 5000):
            chunk = skus[i : i + 5000]
            for product in db.scalars(select(Product).where(Product.sku.in_(chunk))).all():
                existing_by_sku.setdefault(product.sku, product)

        update_fields = (
            "name",
            "ean",
            "brand",
            "category",
            "manufacturer_code",
            "model",
            "own_price",
            "product_url",
            "image_url",
            "description",
            "variant",
            "color",
            "size",
            "storage",
            "memory",
            "supplier_sku",
        )

        if progress_callback:
            progress_callback(0, len(payloads), "importing")
        pending = 0
        applied = 0
        for payload in payloads:
            existing = existing_by_sku.get(payload.sku)
            if existing is not None:
                for field in update_fields:
                    setattr(existing, field, getattr(payload, field))
                existing.import_batch_id = batch.id
            else:
                product = Product(**payload.model_dump(), import_batch_id=batch.id)
                db.add(product)
                existing_by_sku[payload.sku] = product
            imported_rows += 1
            pending += 1
            applied += 1
            if pending >= 2000:
                db.commit()
                pending = 0
                if progress_callback:
                    progress_callback(applied, len(payloads), "importing")
        if pending:
            db.commit()
        if progress_callback:
            progress_callback(applied, len(payloads), "importing")
    finally:
        wb.close()

    skipped_rows = total_rows - imported_rows
    batch.total_rows = total_rows
    batch.imported_rows = imported_rows
    batch.skipped_rows = skipped_rows
    db.commit()
    return ProductImportSummary(
        total_rows=total_rows,
        imported_rows=imported_rows,
        skipped_rows=skipped_rows,
        errors=errors,
    )


def _get_product_by_sku(db: Session, sku: str) -> Product | None:
    stmt = select(Product).where(Product.sku == sku).order_by(Product.created_at.asc()).limit(1)
    return db.scalars(stmt).first()


def build_template_workbook_bytes() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "products"
    for col, h in enumerate(TEMPLATE_HEADERS, start=1):
        ws.cell(row=1, column=col, value=h)
    for col, value in enumerate(TEMPLATE_EXAMPLE_ROW, start=1):
        ws.cell(row=2, column=col, value=value)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
