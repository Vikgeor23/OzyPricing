"""Excel export for competitor workspace rows."""

from __future__ import annotations

import uuid
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import nullslast, or_, select
from sqlalchemy.orm import Session, aliased

from app.db.latest_price import best_match_subquery
from app.models import Competitor, CompetitorCategory, CompetitorProduct, Product
from app.services.workspace_query import (
    WorkspaceQueryParams,
    _category_path_cache,
    _status_filter_condition,
)

MAX_EXPORT_ROWS = 50_000

HEADERS = [
    "Competitor",
    "Title",
    "Category path",
    "URL",
    "Brand",
    "Code",
    "EAN",
    "Manufacturer code",
    "Model",
    "Shop code",
    "Extra code",
    "Size",
    "Final EUR",
    "Regular EUR",
    "Promo EUR",
    "Old/list EUR",
    "Currency",
    "Availability",
    "Offered by",
    "Delivered by",
    "Last checked",
    "Matched SKU",
    "Matched name",
    "Score",
    "Match",
    "Status",
    "Reason",
]


def _text(value: Any) -> str:
    if value is None:
        return ""
    return str(value)


def _attrs_text(attrs: dict[str, Any]) -> str:
    return "; ".join(f"{k}: {v}" for k, v in attrs.items() if k and v is not None)


def _dt_text(value: Any) -> str:
    if value is None:
        return ""
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return str(value)


def _row_values(row) -> list[Any]:
    return [
        row.competitor_name,
        row.title,
        " > ".join(row.category_path or []),
        row.url,
        row.listing_brand,
        row.listing_sku,
        row.listing_ean,
        row.listing_manufacturer_code,
        row.listing_model,
        row.listing_shop_code,
        row.listing_extra_code,
        row.listing_size,
        row.latest_price,
        row.regular_price,
        row.promo_price,
        row.old_price,
        row.currency,
        row.availability,
        row.offered_by,
        row.delivered_by,
        _dt_text(row.last_checked_at),
        row.matched_sku,
        row.matched_product_name,
        row.match_score,
        row.match_method,
        row.match_status,
        row.match_reason,
    ]


def _listing_size(cp: CompetitorProduct) -> str | None:
    return _size_from_json(cp.raw_identifiers, cp.specs_json)


def _size_from_json(raw_identifiers: Any, specs_json: Any) -> str | None:
    raw = raw_identifiers if isinstance(raw_identifiers, dict) else {}
    specs = specs_json if isinstance(specs_json, dict) else {}
    for source in (raw, raw.get("attributes"), specs):
        if not isinstance(source, dict):
            continue
        for key in ("size", "разфасовка", "вместимост", "volume", "capacity"):
            value = source.get(key)
            if value:
                return str(value)
    return None


def _raw_row_values(row, *, competitor_name: str, path_cache: dict[uuid.UUID, list[str]]) -> list[Any]:
    """Build one export row from a lightweight Core result Row (not an ORM
    object — the export selects only the columns it needs and streams them, so
    50k rows never hydrate 50k tracked entities)."""
    category_path = path_cache.get(row.competitor_category_id, []) if row.competitor_category_id else []
    effective_price = row.latest_promo_price if row.latest_promo_price is not None else row.latest_price
    # Match info mirrors the workspace grid: prefer a confirmed direct link, else
    # the best row from product_matches (auto_matched / needs_review / …), so an
    # export shows the same matches the UI does — not only confirmed product_ids.
    display_sku = row.direct_sku or row.match_sku
    display_name = row.direct_name or row.match_name
    match_status = row.match_status_raw or ("confirmed" if row.product_id else "no_candidate")
    return [
        competitor_name,
        row.title,
        " > ".join(category_path),
        row.url,
        row.brand,
        row.sku,
        row.ean,
        row.manufacturer_code,
        row.model,
        row.shop_code,
        row.extra_code,
        _size_from_json(row.raw_identifiers, row.specs_json),
        effective_price,
        row.latest_price,
        row.latest_promo_price,
        row.latest_old_price,
        (row.latest_currency or "BGN") or "BGN",
        row.latest_availability,
        row.latest_offered_by,
        row.latest_delivered_by,
        _dt_text(row.latest_scraped_at or row.last_seen_at),
        display_sku,
        display_name,
        row.match_score,
        row.match_method,
        match_status,
        row.match_reason,
    ]


def _scraped_expr():
    return or_(
        CompetitorProduct.latest_scraped_at.isnot(None),
        CompetitorProduct.latest_price.isnot(None),
        CompetitorProduct.latest_promo_price.isnot(None),
    )


def _apply_fast_filters(stmt, params: WorkspaceQueryParams):
    if params.search:
        pattern = f"%{params.search.strip()}%"
        stmt = stmt.where(
            or_(
                CompetitorProduct.title.ilike(pattern),
                CompetitorProduct.url.ilike(pattern),
                CompetitorProduct.sku.ilike(pattern),
                CompetitorProduct.ean.ilike(pattern),
                CompetitorProduct.brand.ilike(pattern),
                CompetitorProduct.manufacturer_code.ilike(pattern),
                CompetitorProduct.model.ilike(pattern),
                Product.sku.ilike(pattern),
                Product.name.ilike(pattern),
                Product.ean.ilike(pattern),
                Product.brand.ilike(pattern),
                Product.manufacturer_code.ilike(pattern),
                Product.model.ilike(pattern),
            ),
        )
    if params.has_price is True:
        stmt = stmt.where(or_(CompetitorProduct.latest_promo_price.isnot(None), CompetitorProduct.latest_price.isnot(None)))
    elif params.has_price is False:
        stmt = stmt.where(CompetitorProduct.latest_promo_price.is_(None), CompetitorProduct.latest_price.is_(None))
    if params.scraped is True:
        stmt = stmt.where(_scraped_expr())
    elif params.scraped is False:
        stmt = stmt.where(~_scraped_expr())
    if params.status:
        # Full status semantics (auto_matched / needs_review / low_confidence /
        # rejected / confirmed / no_candidate) driven off product_matches — the
        # old two-branch form silently ignored every status but those two.
        stmt = stmt.where(_status_filter_condition(params.status))
    return stmt


def build_workspace_export_xlsx(
    db: Session,
    *,
    competitor_id: uuid.UUID,
    category_id: uuid.UUID | None,
    params: WorkspaceQueryParams,
) -> bytes | None:
    competitor = db.get(Competitor, competitor_id)
    if competitor is None:
        return None
    if category_id is not None:
        category = db.get(CompetitorCategory, category_id)
        if category is None or category.competitor_id != competitor_id:
            return None

    # Write-only mode streams rows straight to the on-disk XLSX parts instead of
    # holding a Cell object per cell in memory. For a 50k-row export that is the
    # difference between ~530 MB / 80 s and a small, near-constant footprint.
    wb = Workbook(write_only=True)
    ws = wb.create_sheet("Competitors")
    ws.freeze_panes = "A2"
    widths = [18, 42, 28, 52, 18, 18, 18, 20, 18, 16, 16, 12, 36, 60, 14, 14, 14, 14, 10, 16, 24, 18, 32, 12, 16, 16, 36]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    header_fill = PatternFill("solid", fgColor="111827")
    header_font = Font(color="FFFFFF", bold=True)
    header_align = Alignment(wrap_text=True, vertical="top")
    header_cells = []
    for header in HEADERS:
        cell = WriteOnlyCell(ws, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        header_cells.append(cell)
    ws.append(header_cells)

    # Data cells are appended as bare Python values (no per-cell style objects):
    # wrapping every one of ~1.35M cells in a styled WriteOnlyCell cost ~40 s and
    # doubled the file with no data benefit — column widths carry the layout.

    scope_where = (
        CompetitorProduct.competitor_category_id == category_id
        if category_id is not None
        else CompetitorProduct.competitor_id == competitor_id
    )
    # Same match resolution as the workspace grid: `best` picks one row per
    # listing from product_matches (confirmed > auto_matched > needs_review > …);
    # Product (direct) is the confirmed link, pm_product the best-match product.
    # The window is scoped to this competitor's listings — the export streams
    # every row, so an unscoped global window would re-rank all matches per run.
    cp_scope = select(CompetitorProduct.id).where(scope_where)
    best = best_match_subquery(include_rejected=True, cp_scope=cp_scope)
    pm_product = aliased(Product)
    stmt = (
        select(
            CompetitorProduct.title,
            CompetitorProduct.url,
            CompetitorProduct.brand,
            CompetitorProduct.sku,
            CompetitorProduct.ean,
            CompetitorProduct.manufacturer_code,
            CompetitorProduct.model,
            CompetitorProduct.shop_code,
            CompetitorProduct.extra_code,
            CompetitorProduct.raw_identifiers,
            CompetitorProduct.specs_json,
            CompetitorProduct.competitor_category_id,
            CompetitorProduct.product_id,
            CompetitorProduct.latest_price,
            CompetitorProduct.latest_promo_price,
            CompetitorProduct.latest_old_price,
            CompetitorProduct.latest_currency,
            CompetitorProduct.latest_availability,
            CompetitorProduct.latest_offered_by,
            CompetitorProduct.latest_delivered_by,
            CompetitorProduct.latest_scraped_at,
            CompetitorProduct.last_seen_at,
            Product.sku.label("direct_sku"),
            Product.name.label("direct_name"),
            pm_product.sku.label("match_sku"),
            pm_product.name.label("match_name"),
            best.c.match_score.label("match_score"),
            best.c.match_method.label("match_method"),
            best.c.status.label("match_status_raw"),
            best.c.match_reason.label("match_reason"),
        )
        .outerjoin(Product, Product.id == CompetitorProduct.product_id)
        .outerjoin(best, best.c.competitor_product_id == CompetitorProduct.id)
        .outerjoin(pm_product, pm_product.id == best.c.product_id)
        .where(scope_where)
    )
    stmt = (
        _apply_fast_filters(stmt, params)
        .order_by(
            nullslast(CompetitorProduct.latest_scraped_at.desc()),
            CompetitorProduct.created_at.desc(),
        )
        .limit(MAX_EXPORT_ROWS)
    )
    path_cache = _category_path_cache(db, competitor_id)

    # Fetch the whole (already LIMIT-capped) result in one statement: the rows are
    # lightweight Core tuples, not ORM entities. A server-side streaming cursor
    # was tried here but it disables Postgres parallel query — turning a 0.1s plan
    # into 26s — for a memory saving that openpyxl's write-only sheet already
    # provides, so a plain buffered fetch is both faster and within the timeout.
    exported = 0
    for raw_row in db.execute(stmt).all():
        ws.append(_raw_row_values(raw_row, competitor_name=competitor.name, path_cache=path_cache))
        exported += 1

    last_col = get_column_letter(len(HEADERS))
    ws.auto_filter.ref = f"A1:{last_col}{exported + 1}"

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()
